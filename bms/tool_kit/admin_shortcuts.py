from decimal import Decimal
from django.utils import timezone
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from ..models import *


def if_value_none(value):
	if value is not None and not value == '':
		return Decimal(value)
	else:
		return 0

def import_options_rate(self, request, obj, change):
	try:
		wb = load_workbook(filename=obj.file.path)
		sheet_names = wb.sheetnames
		sheet = wb[sheet_names[0]]
		cell_header = list(sheet.rows)[0]
		list_header=[]
		# -------------------------去除为空的列头------------------------
		for header in cell_header:
			if header._value is not None:
				list_header.append(header._value)
		# -------------------------去除为空的列头------------------------

		# -------------------------根据非空列头获取单元格数据------------------------
		lists = []
		for row in range(2,sheet.max_row+1):
			r = {}
			for col in range(1,len(list_header)+1):
				key = list_header[col - 1]
				r[key] = sheet.cell(row=row, column=col).value
			lists.append(r)
		# -------------------------根据非空列头获取单元格数据------------------------

		# -------------------------将数据插入数据库------------------------
		add_list = []
		for cell in lists:
			new_rate = Option_Rate(code=cell[list_header[0]],
			                       name = cell[list_header[1]],
			                       options_file=obj,
			                       op_type = obj.op_type,
			                       call_put=cell[list_header[2]],
			                       day1 =Decimal(if_value_none(cell[list_header[3]])),
			                       day2=Decimal(if_value_none(cell[list_header[4]])),
			                       day3=Decimal(if_value_none(cell[list_header[5]])),
			                       day4=Decimal(if_value_none(cell[list_header[6]])),
			                       day5=Decimal(if_value_none(cell[list_header[7]])),
			                       day6=Decimal(if_value_none(cell[list_header[8]])),
			                       week1=Decimal(if_value_none(cell[list_header[9]])),
			                       week2=Decimal(if_value_none(cell[list_header[10]])),
			                       week3=Decimal(if_value_none(cell[list_header[11]])),
			                       month1=Decimal(if_value_none(cell[list_header[12]])),
	                               month2=Decimal(if_value_none(cell[list_header[13]])),
	                               month3=Decimal(if_value_none(cell[list_header[14]])),
	                               month4=Decimal(if_value_none(cell[list_header[15]])),
	                               month5=Decimal(if_value_none(cell[list_header[16]])),
	                               month6=Decimal(if_value_none(cell[list_header[17]])),
	                               month7=Decimal(if_value_none(cell[list_header[18]])),
	                               month8=Decimal(if_value_none(cell[list_header[19]])),
	                               month9=Decimal(if_value_none(cell[list_header[20]])),
	                               month10=Decimal(if_value_none(cell[list_header[21]])),
	                               month11=Decimal(if_value_none(cell[list_header[22]])),
	                               month12=Decimal(if_value_none(cell[list_header[23]])),
	                               notional_principal = Decimal(if_value_none(cell[list_header[24]])),
	                               operator = request.user.username,
								   effective_date = obj.effective_date
			                       )

			add_list.append(new_rate)

		end = Option_Rate.objects.bulk_create(add_list)
		# -------------------------将数据插入数据库------------------------
		if end is not None and not len(end) == 0:
			return True, ''
		else:
			return False,'期权费率表没有添加数据。'
	except Exception as ex:
		print(ex)
		return False,ex.__str__()
